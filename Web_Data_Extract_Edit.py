from urllib.request import urlopen
from openpyxl import load_workbook
wb = load_workbook("data_update.xlsx")
ws = wb.active
wb1 = load_workbook("Company_Codes.xlsx")
ws1 = wb1.active
wc1 = ws1["C1:C1000"]
	
mainlink = "https://www.moneycontrol.com/financials/results/quarterly-results/"
for ww in range(len(wc1)):
    wwe = wc1[ww]
    wwe1 = wwe[0]
    mainlink2 = mainlink+wwe1.value
    f = urlopen(mainlink2)
    myfile = str(f.read())
    print(wwe1.value)

	wc2 = ws.cell(row=1+ww*8 , column = 3) 
    if wc2.value == 0:
    	print('empty')
    	continue
    for k in range(9,0,-1):
        f2 = urlopen(mainlink2+'/'+str(k))
        myfile2 = str(f2.read())   
        a3 = myfile2.split('</tr>')
        if len(a3) < 3: 
            print('empty page')
            continue
        a4 = a3[2]
        a4count = a4.count('<td>')
        if a4count == 6:
        	print('alright')
        	break
        sum = 0
    #Revenue&EBIDTA
        
        for j in [4,11,21,25,29,30]:
            b3 = a3[j]
            c3 = b3.split('<td>')
            sum = sum +1
            for i in range(1,7-6+a4count):
                d3 = c3[i]
                e3 = d3.split('</td>')
                f3 = e3[0]

                if k == 1:
                    if i != 1:
                        f3 =f3.replace(",","")
                        f3 =f3.replace("--","0")
                        f3 = float(f3)
                    wc = ws.cell(row=1+sum+ww*8 , column = 2+i) 
                    wc.value = f3
                else:
                    if i != 1:
                        f3 =f3.replace(",","")
                        f3 =f3.replace("--","0")
                        f3 = float(f3)
                        wc = ws.cell(row=1+sum+ww*8 , column = 2+i+(5*(k-1))) 
                        wc.value = f3

    print('updated')    
    wb.save('data_update.xlsx')
print('done')    
wb.save('data_update.xlsx')
