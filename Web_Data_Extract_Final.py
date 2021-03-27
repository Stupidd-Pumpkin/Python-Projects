from urllib.request import urlopen
from openpyxl import load_workbook
wb = load_workbook("data3.xlsx")
ws = wb.active
wb1 = load_workbook("Company_Codes3.xlsx")
ws1 = wb1.active
wc1 = ws1["C7961:C8703"]
	
mainlink = "https://www.moneycontrol.com/financials/results/quarterly-results/"
for ww in range(len(wc1)):
    wwe = wc1[ww]
    wwe1 = wwe[0]
    mainlink2 = mainlink+wwe1.value
    f = urlopen(mainlink2)
    myfile = str(f.read())
    print(wwe1.value)

    #CompanyName
    x = myfile.split('B hidden-xs">')
    y = x[1]
    w = y.split('<',1)
    z = w[0]
    for i in range(0,7):
        wc = ws.cell(row=1+i+ww*8 , column = 1) 
        wc.value = z

    #SectorName
    x2 = myfile.split('lg">',1)
    y2 = x2[1]
    w2 = y2.split('</span',1)
    z2 = w2[0]
    for i in range(0,7):
        wc = ws.cell(row=1+i++ww*8 , column = 2) 
        wc.value = z2
    if z2 == 'Banks - Public Sector':
        continue
    if z2 == 'Banks - Private Sector':
        continue
    
    #HeadingName
    a = myfile.split('<td>')
    if len(a) < 3:
        continue
    b = a[2]
    c = b.split('<span')
    d = c[0]
    wc = ws.cell(row=1+ww*8 , column = 3) 
    wc.value = d
    
    for k in range(1,10):
        f2 = urlopen(mainlink2+'/'+str(k))
        myfile2 = str(f2.read())
        h = myfile2.split('<td>')
        
        a3 = myfile2.split('</tr>')
        if len(a3) < 3: 
            continue
        a4 = a3[2]
        a4count = a4.count('<td>')
        

        #HeadingPeriod
        for i in range(3,8-6+a4count):
            a2 = h[i]
            b2 = a2.split('</td>')
            c2 = b2[0]
            d2 = c2.replace(' &#039;','-')
            wc = ws.cell(row=1+ww*8 , column = 1+i+(5*(k-1))) 
            wc.value = d2

        sum = 0
    #Revenue&EBIDTA
        
        for j in [4,11,21,25,29,30]:
            b3 = a3[j]
            c3 = b3.split('<td>')
            sum = sum +1
            for i in range(1,7-6+a4count):
                d3 = c3[i]
                e3 = d3.split('</td>')
                f3 = e3[0]

                if k == 1:
                    if i != 1:
                        f3 =f3.replace(",","")
                        f3 =f3.replace("--","0")
                        f3 = float(f3)
                    wc = ws.cell(row=1+sum+ww*8 , column = 2+i) 
                    wc.value = f3
                else:
                    if i != 1:
                        f3 =f3.replace(",","")
                        f3 =f3.replace("--","0")
                        f3 = float(f3)
                        wc = ws.cell(row=1+sum+ww*8 , column = 2+i+(5*(k-1))) 
                        wc.value = f3

        
    wb.save('data3.xlsx')
print('done')    

wb.save('data3.xlsx')