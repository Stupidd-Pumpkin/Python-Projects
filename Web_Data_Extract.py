from urllib.request import urlopen
from openpyxl import Workbook
#import winsound
#frequency = 2500  # Set Frequency To 2500 Hertz
#duration = 1000  # Set Duration To 1000 ms == 1 second

link3 = "https://www.moneycontrol.com/stocks/sectors/food-processing.html"
f = urlopen(link3)
myfile = str(f.read())
data3 = myfile.split("<span></span> <a href=\"http://www.moneycontrol.com/stocks/sectors/")
data3 = data3[1:]
L3 = len(data3)
for i in range(L3):
    y = data3[i]
    z = y.split(".")
    data3[i] = z[0]
#print(data3)

mainlink2 = "https://www.moneycontrol.com/stocks/sectors/"

for n in range(8,L3):
    wb_name = data3[n]+".xlsx"
    wb = Workbook()
    link2 = mainlink2+data3[n]+".html"
    print(link2)
    f = urlopen(link2)
    myfile = str(f.read())
    data2 = myfile.split("<td class=\"left\"><a href=")
    data2 = data2[1:]
    #y = data[i].split("")
    L2 = len(data2)
    for i in range(L2):
        y = data2[i]
        z = y.split("\">")
        z = z[0]
        w = z.split("/")
        w = w[-1]
        #print(w)
        data2[i] = w
    print(data2)

    for m in range(L2):
        #example for a mainlink "https://www.moneycontrol.com/financials/results/quarterly-results/PGH"
        mainlink = "https://www.moneycontrol.com/financials/results/quarterly-results/"
        #print(mainlink)
        #sheet = wb.create_sheet(title = "Procter and Gamble Hygiene and Health Care Ltd.")
        sheet = wb.create_sheet(title = data2[m])

        k = 0
        for l in range(1,11):
            link = mainlink+data2[m]+"/"+str(l)
            print(link)
            f = urlopen(link)
            myfile = str(f.read())
            data = myfile.split("<td>")
            L = len(data)
            for i in range(L):
                y = data[i]
                z = y.split("<")
                z = z[0]
                z = z.replace("&#039;","")
                data[i] = z
            max=L
            for i in range(L):
                if data[i] == 'Public Share Holding':
                    max = i
                    break
            data = data[2:max]        
            L = len(data)
            n = 6
            if L < 38*2: break
            if L < 38*3: n = 2
            if L < 38*4: n = 3
            if L < 38*5: n = 4
            if L < 38*6: n = 5
            #print(n)    
            for i in range(int(L/n)):
                for j in range(n):
                    if (j != 0):
                        x = sheet.cell(row= i+1 , column = (l-1)*6+j-k+1) 
                        x.value = data[i*n+j]
                    elif(l == 1):
                        x = sheet.cell(row= i+1 , column = (l-1)*6+j-k+1) 
                        x.value = data[i*n+j]
            k = k+1;
        wb.save(wb_name)
        #winsound.Beep(frequency, duration)